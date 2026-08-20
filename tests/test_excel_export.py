from io import BytesIO
from openpyxl import load_workbook
from superbid_collector.storage import Store
from superbid_collector.excel_export import export_operational_workbook
def test_export_workbook(tmp_path):
 s=Store(tmp_path/"x.db");s.init();wb=load_workbook(BytesIO(export_operational_workbook(s.conn)),read_only=True);assert "Subastas activas" in wb.sheetnames;assert "Historico" in wb.sheetnames;assert "Peritajes y anexos" in wb.sheetnames;assert "Operacion collector" in wb.sheetnames
