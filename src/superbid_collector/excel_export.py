from __future__ import annotations
from io import BytesIO
import sqlite3
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from .history import historical_rows
from .dashboard_service import active_opportunities
from .settings import get_cost_profile

def _sheet(ws,headers,rows):
    ws.append(headers)
    for row in rows:ws.append([row.get(h) for h in headers])
    ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
    for i,h in enumerate(headers,1):
        max_len=max([len(str(h))]+[len(str(ws.cell(r,i).value or "")) for r in range(2,min(ws.max_row,250)+1)]); ws.column_dimensions[get_column_letter(i)].width=min(max(max_len+2,10),45)

def export_operational_workbook(conn:sqlite3.Connection)->bytes:
    wb=Workbook(); wb.remove(wb.active); profile=get_cost_profile(conn); active=active_opportunities(conn,profile,10000); hist=historical_rows(conn,limit=100000)
    atts=[dict(r) for r in conn.execute("SELECT l.external_lot_id,l.title,a.kind,a.name,a.url,a.discovered_at FROM lot_attachments a JOIN lots l ON l.id=a.lot_id ORDER BY a.discovered_at DESC").fetchall()]
    runs=[dict(r) for r in conn.execute("SELECT run_type,target,started_at,finished_at,ok,lots_found,lots_saved,attachments_saved,bids_saved,error FROM collection_runs ORDER BY started_at DESC LIMIT 5000").fetchall()]
    ws=wb.create_sheet("Subastas activas"); _sheet(ws,["external_lot_id","title","brand","model_year","city","seller","current_bid_cop","closes_at_text","bid_count","peritaje_available","historical_reference_cop","market_reference_cop","max_bid_cop","expected_profit_cop","expected_roi_pct","score","decision","url"],active)
    ws=wb.create_sheet("Historico"); _sheet(ws,["external_lot_id","title","brand","model_year","city","seller","initial_bid_cop","outcome","closing_price_observed_cop","sale_price_confirmed_cop","historical_value_cop","confidence","last_seen_at","data_quality","data_confidence","data_source_type","url"],hist)
    ws=wb.create_sheet("Peritajes y anexos"); _sheet(ws,["external_lot_id","title","kind","name","url","discovered_at"],atts)
    prov=[dict(r) for r in conn.execute("SELECT l.external_lot_id,l.title,p.source_type,p.source_url,p.observed_at,p.confidence,p.note FROM lot_provenance p JOIN lots l ON l.id=p.lot_id ORDER BY p.observed_at DESC").fetchall()]
    ws=wb.create_sheet("Proveniencia"); _sheet(ws,["external_lot_id","title","source_type","source_url","observed_at","confidence","note"],prov)
    ws=wb.create_sheet("Operacion collector"); _sheet(ws,["run_type","target","started_at","finished_at","ok","lots_found","lots_saved","attachments_saved","bids_saved","error"],runs)
    ws=wb.create_sheet("Parametros")
    for k,v in profile.__dict__.items():ws.append([k,v])
    ws.column_dimensions["A"].width=35; ws.column_dimensions["B"].width=20
    out=BytesIO(); wb.save(out); return out.getvalue()
