from __future__ import annotations

import hashlib
import re
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from .market_storage import init_market_schema


BASE_NAMES = {
    "codigo": ["codigo", "código", "codigo activo", "código activo"],
    "homologo": ["homologoco", "homologo", "homólogo", "codigo homologo", "código homólogo"],
    "marca": ["marca"],
    "clase": ["clase"],
    "ref1": ["referencia1", "referencia 1", "ref1"],
    "ref2": ["referencia2", "referencia 2", "ref2"],
    "ref3": ["referencia3", "referencia 3", "ref3"],
    "servicio": ["servicio"],
}


def _norm(v: Any) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip().lower()



def fasecolda_record_key(
    *, source_file: str, code: str, homologous_code: str, brand: str,
    vehicle_class: str, reference1: str, reference2: str, reference3: str,
    service: str, model_year: int,
) -> str:
    parts = [
        source_file, code, homologous_code, brand, vehicle_class,
        reference1, reference2, reference3, service, str(model_year),
    ]
    canonical = "|".join(_norm(x) for x in parts)
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()


def _find_header(ws, max_scan=60):
    for row_idx in range(1, min(ws.max_row, max_scan) + 1):
        vals = [_norm(c.value) for c in ws[row_idx]]
        has_brand = any(v == "marca" for v in vals)
        year_count = sum(1 for v in vals if re.fullmatch(r"(?:19|20)\d{2}", v))
        if has_brand and year_count >= 3:
            return row_idx
    raise ValueError("No se encontró encabezado Fasecolda reconocible.")


def _colmap(ws, header_row):
    out = {}
    for cell in ws[header_row]:
        n = _norm(cell.value)
        if not n:
            continue
        if re.fullmatch(r"(?:19|20)\d{2}", n):
            out[f"year:{n}"] = cell.column
            continue
        for key, options in BASE_NAMES.items():
            if n in options:
                out[key] = cell.column
    if "marca" not in out:
        raise ValueError("Falta columna Marca.")
    return out


def _value_cop(v: Any) -> int | None:
    if v in (None, "", 0, "0"):
        return None
    try:
        x = float(str(v).replace(",", "").replace("$", "").strip())
    except Exception:
        return None
    if x <= 0:
        return None
    # Fasecolda workbooks commonly express guide values in thousands of COP.
    # If already in full pesos, preserve it.
    return int(round(x * 1000 if x < 1_000_000 else x))


def import_fasecolda_excel(conn: sqlite3.Connection, path: str | Path) -> dict:
    init_market_schema(conn)
    path = Path(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    inserted = 0
    sheets = 0
    imported_at = datetime.now(timezone.utc).isoformat()

    for ws in wb.worksheets:
        try:
            h = _find_header(ws)
            cmap = _colmap(ws, h)
        except ValueError:
            continue
        sheets += 1

        year_cols = {
            int(k.split(":")[1]): col
            for k, col in cmap.items()
            if k.startswith("year:")
        }
        for row_idx in range(h + 1, ws.max_row + 1):
            def val(key):
                col = cmap.get(key)
                return ws.cell(row=row_idx, column=col).value if col else None

            brand = val("marca")
            if not brand:
                continue

            for year, col in year_cols.items():
                amount = _value_cop(ws.cell(row=row_idx, column=col).value)
                if amount is None:
                    continue
                row_data = {
                    "source_file": path.name,
                    "code": str(val("codigo") or ""),
                    "homologous_code": str(val("homologo") or ""),
                    "brand": str(brand),
                    "vehicle_class": str(val("clase") or ""),
                    "reference1": str(val("ref1") or ""),
                    "reference2": str(val("ref2") or ""),
                    "reference3": str(val("ref3") or ""),
                    "service": str(val("servicio") or ""),
                    "model_year": year,
                }
                record_key = fasecolda_record_key(**row_data)
                cur = conn.execute(
                    """
                    INSERT INTO fasecolda_values (
                      record_key, source_file, imported_at, code, homologous_code, brand,
                      vehicle_class, reference1, reference2, reference3,
                      service, model_year, value_cop
                    ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
                    ON CONFLICT(record_key) DO UPDATE SET
                      imported_at=excluded.imported_at,
                      value_cop=excluded.value_cop
                    """,
                    (
                        record_key, row_data["source_file"], imported_at, row_data["code"],
                        row_data["homologous_code"], row_data["brand"], row_data["vehicle_class"],
                        row_data["reference1"], row_data["reference2"], row_data["reference3"],
                        row_data["service"], year, amount,
                    ),
                )
                inserted += 1

    conn.commit()
    return {"source_file": path.name, "sheets": sheets, "values_inserted": inserted}
